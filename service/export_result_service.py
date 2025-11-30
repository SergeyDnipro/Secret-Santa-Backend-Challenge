from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def xls_generate(game_data, filename:str):
    wb = Workbook()
    ws = wb.active

    ws.title = game_data["game"][2]

    headers = ["ID", "Player name", "Receiver name"]
    ws.append(headers)

    # Fill row with player data
    for player in game_data["players"]:
        row = [
            player[0],
            player[1],
            player[3]
        ]
        ws.append(row)

    # Auto-width
    for i, col in enumerate(ws.columns, start=1):
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        ws.column_dimensions[get_column_letter(i)].width = max_length + 2

    wb.save(filename)


def export_xls_worker(bot, chat_id, game_data):
    """ Background thread to generate result file """

    # Create filename string
    date, time = game_data['game'][1].split()
    time = time.split(":")
    date_time = '__'.join([date, '_'.join(time)])
    filename = f"{game_data['game'][2]}_{date_time}.xlsx"

    # Generate XLS document
    xls_generate(game_data, filename)
    with open(filename, "rb") as f:
        bot.send_document(chat_id, f)